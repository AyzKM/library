from flask import Flask, render_template, request
from openpyxl import load_workbook
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.orm import scoped_session
from database import engine, Book

app = Flask(__name__)

db = scoped_session(sessionmaker(bind=engine))

@app.route("/")
def homepage():
    if 'key_word' in request.args:
        key_word = request.args.get("key_word")
        session = sessionmaker(engine)()
        books = session.execute(f"""
            SELECT * FROM "Book"
            WHERE name LIKE '%{key_word}%'
            OR author LIKE '%{key_word}%'
            ;
        """)
        session.commit()
    else:
        with engine.connect() as con:
            books = con.execute("""SELECT * FROM "Book";""")

    return render_template("homepage.html", object_list=books)

# ver with sesionmaker
    # session = sessionmaker(bind=engine)()
    # books = session.query(Book)
    # print(books)
    # session.commit()

    # return render_template("homepage.html", object_list=books)

#old version with excel
# @app.route("/")
# def homepage():
#     excel = load_workbook("tales.xlsx")
#     page = excel["Sheet"]
#     object_list = [[tale.value, tale.offset(column=1).value] for tale in page["A"][1:]]
#     return render_template("homepage.html", object_list=object_list)


@app.route("/books/")
def books():
# v_4 with raw query
    with engine.connect() as con:
        books = con.execute("""SELECT * FROM "Book";""")

    return render_template("books.html", object_list=books)

#v_3 with sessionmaker
#     # session = sessionmaker(engine)()
#     # books = session.query(Book)
#     # session.commit()   
#     # print(books)

#v_2 with excel
# #     excel = load_workbook("tales.xlsx")
# #     page = excel["Sheet"]

# #     object_list = [[tale.value, tale.offset(column=1).value] for tale in page["A"][1:]]
# #     print(object_list)
# #     return render_template(
# #         "books.html", object_list=object_list
# #         )

# v_1
# #     tales = [tale.value for tale in page["A"]][1:]

# #     authors = [author.value for author in page["B"]][1:]
     
# #     html = """
# #         <a href="/authors">Авторы</a>
# #         <a href="/books">Книги</a>
# #         <h1>Тут будет список книг</h1>
# #     """

# #     for i in range(len(tales)):
# #         html += f"<h2>{tales[i]} - {authors[i]}</h2>"

# #     return html 

@app.route("/authors/")
def authors():
# ver with sesionmaker
    session = sessionmaker(bind=engine)()
    books = session.query(Book)
    session.commit()

    return render_template("authors.html", authors=books)

#  ver with raw request
# @app.route("/db/authors/")
# def db_authors():
#     with engine.connect() as con:
#         authors = con.execute('SELECT DISTINCT author FROM "Book";')

#     return render_template("authors.html", authors=authors)

# # old version with excel
# @app.route("/authors/")
# def authors():
#     excel = load_workbook("tales.xlsx")
#     page = excel["Sheet"]
#     authors = {author.value for author in page["B"][1:]}
#     return render_template("authors.html", authors=list(authors))


@app.route("/form/")
def form():
    return render_template("form.html")


@app.route("/add/", methods=["POST"])
def book_add():
# ver with raw request
    f = request.form
    book = f["book"]
    author = f["author"]
    # url = f["url"]

    ids = db.execute('SELECT id FROM "Book" ORDER BY id DESC;')
    max_id = ids.first().id
    c_id = max_id + 1

    db.execute(f'''
        INSERT INTO "Book" (id, name, author)
        VALUES ({c_id}, '{book}', '{author}');   
    ''')

    db.commit()

    return render_template("saved.html")

# old ver with excel (did not work)
# @app.route("/add/", methods=["POST"])
# def book_add():
#         f = request.form 
#         # print(f["Book"], f["Author"])
#         excel = load_workbook("tales.xlsx")
#         page = excel["Sheet"]
#         last = len(page["A"]) + 1
#         page[f"A{last}"] = f["Book"]
#         page[f"B{last}"] = f["Author"]
#         excel.save("tales.xlsx")
#         return render_template("book_add.html")

# @app.route("/add/save/")
# def book_new_save():
#     return render_template("saved.html")


@app.route("/book/<id>/")
def book(id):
    obj = db.execute(f'SELECT * FROM "Book" WHERE id = {id};').first()
    return render_template("book.html", obj=obj)


# old verion with excel
    # excel = load_workbook("tales.xlsx")
    # page = excel["Sheet"]
    # object_list = [[tale.value, tale.offset(column=1).value, tale.offset(column=2).value] for tale in page["A"][1:]]
    # obj = object_list[int(num)]
    # obj.append(num)
    # return render_template(
    #     "book.html", obj=obj
    #     )

# @app.route("/<int:id>/", methods=["GET","POST"])
# def book_edit(id):
#     if request.method == "POST":
#         name = request.form.get("book")
#         author = request.form.get("author")
#         # image = request.form.get("image")
#         db.execute(f'''
#             UPDATE "Book"
#             SET
#             name = '{name}',
#             author = '{author}'
#         WHERE id = {id};
#         ''')
#         db.commit()
#     else:
#         return 'method get'

#     book_object = db.execute(f'SELECT * FROM "Book" WHERE id = {id};').first()
#     return render_template("book_edit.html", book_object=book_object)


@app.route("/<int:id>/edit")
def book_edit_form(id):
    book_object = db.execute(f'SELECT * FROM "Book" WHERE id = {id};').first()
    return render_template("book_edit.html", book_object = book_object)


@app.route("/<int:id>/", methods=["POST"])
def book_edit(id):
    name = request.form.get("book")
    author = request.form.get("author")
        # image = request.form.get("image")
    db.execute(f'''
        UPDATE "Book"
        SET
        name = '{name}',
        author = '{author}'
        WHERE id = {id};
    ''')
    db.commit()

    return render_template("saved.html")

# @app.route("/book/<num>/edit/")
# def book_edit(num):

# # old version with excel
#     num = int(num) + 2
#     excel = load_workbook("tales.xlsx")
#     page = excel["Sheet"]
#     tale = page[f"A{num}"]
#     author = page[f"B{num}"]
#     image = page[f"C{num}"]
#     obj = [tale.value, author.value, image.value, num]
#     return render_template("book_edit.html", obj=obj)


# @app.route("/book/<num>/save/", methods=["POST", "GET"])
# def book_save(num):
#     if request.method == "POST":
#         num = int(num)
#         excel = load_workbook("tales.xlsx")
#         page = excel["Sheet"]
#         form = request.form
#         page[f"A{num}"] = form["tale"]
#         page[f"B{num}"] = form["author"]
#         page[f"C{num}"] = form["image"]
#         excel.save("tales.xlsx")
#         return render_template("saved.html")
#     else:
#         print("this method is get")

